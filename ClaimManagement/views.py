from openpyxl import load_workbook
from django.shortcuts import render
from openpyxl.styles import Font
import xlrd
from openpyxl.utils import get_column_letter
import openpyxl
from rest_framework import generics
from .serializers import *
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from .models import *
from django.core.files.uploadedfile import SimpleUploadedFile
import os
import datetime
import zipfile
import pandas as pd
import io
from rest_framework import status
from PreAuth.views import PDFGenerator
# Create your views here.

# This is for .CSV files


class ExcelMergeCSVAPIView(generics.GenericAPIView):
    parser_classes = [MultiPartParser]
    serializer_class = ExcelFileSerializer

    def post(self, request):
        serializer = ExcelFileSerializer(data=request.data)
        if serializer.is_valid():
            # Load the CSV files
            csv_files = request.FILES.getlist('file')
            if len(csv_files) < 2:
                return Response({'message': 'At least two CSV files are required for Merging.',
                                 'status': 'error'},
                                status=status.HTTP_400_BAD_REQUEST)

            # Merge the CSV files
            dfs = [pd.read_csv(f, header=2) for f in csv_files]
            merged_df = pd.concat(dfs, ignore_index=True)

            # Write the merged data to an XLSX file
            output_file = io.BytesIO()
            writer = pd.ExcelWriter(output_file, engine='openpyxl')
            merged_df.to_excel(writer, index=False, sheet_name='Sheet1')

            # Add the header only once
            worksheet = writer.sheets['Sheet1']
            for i, col in enumerate(merged_df.columns):
                worksheet.cell(row=1, column=i+1, value=col)

            # Save the XLSX file
            writer.save()
            output_file.seek(0)
            date = datetime.datetime.today().date()
            # Return the merged file as a response
            response = HttpResponse(output_file.read(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename= MERGE_DUMP_{date}.xlsx'.format(
                date=date)
            return response
# This is for .xls files


class ExcelMergexlsAPIView(generics.GenericAPIView):
    serializer_class = ExcelFileSerializer
    parser_classes = [MultiPartParser]

    def post(self, request, format=None):
        serializer = ExcelFileSerializer(data=request.data)
        if serializer.is_valid():
            # Load the Excel files
            excel_files = []
            for file in request.FILES.getlist('file'):
                excel_files.append(xlrd.open_workbook(
                    file_contents=file.read()))
            if len(excel_files) == 1:
                return Response({'status': 'error',
                                'message': "There is only one Excel file ,\
                                         To Merge the Excel's at least two or more files will be needed."})
            # Merge the Excel files

            merged_workbook = openpyxl.Workbook()
            merged_sheet = merged_workbook.active
            merged_sheet.title = 'Merged Sheet'
            header_added = False  # To keep track of whether the header has been added
            row_index_merged = 1  # To keep track of the row index in the merged sheet
            for excel_file in excel_files:
                for sheet in excel_file.sheets():
                    for row_index in range(sheet.nrows):
                        # If the sheet is the first sheet, and header has not been added yet
                        if not header_added and row_index == 2:
                            # Add the headers to the merged sheet
                            for col_index in range(sheet.ncols):
                                column_letter = get_column_letter(col_index+1)
                                merged_sheet[column_letter+str(row_index_merged)] = sheet.cell_value(
                                    row_index, col_index)
                                merged_sheet[column_letter +
                                             str(row_index_merged)].font = Font(bold=True)
                            row_index_merged += 1
                            header_added = True  # Set header_added to True so that it is not added again
                        elif row_index > 2:  # If the row is after the header row
                            for col_index in range(sheet.ncols):
                                column_letter = get_column_letter(col_index+1)
                                merged_sheet[column_letter+str(row_index_merged)] = sheet.cell_value(
                                    row_index, col_index)
                            row_index_merged += 1

            # Write the merged Excel file to a BytesIO object
            response = HttpResponse(
                content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=MERGE_DUMP.xlsx'
            merged_workbook.save(response)

            return response
        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'})


class CaseSheetPostForm(generics.GenericAPIView):
    serializer_class = CaseSheetSerializer
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        CaseNumber = request.data['CaseNumber'].replace('/', '')
        date = datetime.datetime.today().date()
        files = {}
        for field in ['AdmitCase_ClinicalSheet', 'ICP', 'MedicationChart', 'InitialSheet', 'TPRChart',
                      'VitalChart', 'LOS_StayChart', 'CaseSheets_OtheDocuments']:

            images = request.FILES.getlist(field)
            if images:
                for image in images:
                    if image.size > 400 * 1024:
                        return Response({'status': 'error', 'message': 'Each Image size should be less than 499 KB'}, status=400)
                pdf = PDFGenerator(images, CaseNumber, field)
                files[field] = SimpleUploadedFile(
                    f"{CaseNumber}_{date}_{field}.zip", pdf, content_type='application/zip')

        serializer = self.get_serializer(data={
            'CaseNumberID': request.data['CaseNumberID'],
            'CaseNumber': request.data['CaseNumber'],
            'AdmitCase_ClinicalSheet': files.get('AdmitCase_ClinicalSheet', None),
            'ICP': files.get('ICP', None),
            'MedicationChart': files.get('MedicationChart', None),
            'InitialSheet': files.get('InitialSheet', None),
            'TPRChart': files.get('TPRChart', None),
            'VitalChart': files.get('VitalChart', None),
            'LOS_StayChart': files.get('LOS_StayChart', None),
            'CaseSheets_OtheDocuments': files.get('CaseSheets_OtheDocuments', None), })
        # print(serializer.pop('CaseNumber'))

        if serializer.is_valid():
            CaseNumber = serializer.validated_data.pop('CaseNumber')
            serializer.save(user=request.user)
            return Response({'status': 'success',
                            'message': 'data saved successfully'})

        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'}, status=400)


class LabtestPostForm(generics.GenericAPIView):
    serializer_class = LabTestSerializer
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        CaseNumber = request.data['CaseNumber'].replace('/', '')
        date = datetime.datetime.today().date()
        files = {}
        for field in ['Microbiology', 'Biochemistry', 'Pathology', 'SerologyInvestigation', 'OtherDocuments']:

            images = request.FILES.getlist(field)
            if images:
                for image in images:
                    if image.size > 400 * 1024:
                        return Response({'status': 'error', 'message': 'Each Image size should be less than 499 KB'}, status=400)
                pdf = PDFGenerator(images, CaseNumber, field)
                files[field] = SimpleUploadedFile(
                    f"{CaseNumber}_{date}_{field}.zip", pdf, content_type='application/zip')

        serializer = LabTestSerializer(data={
            'CaseNumberID': request.data['CaseNumberID'],
            'CaseNumber': request.data['CaseNumber'],
            'Microbiology': files.get('Microbiology', None),
            'Biochemistry': files.get('Biochemistry', None),
            'Pathology': files.get('Pathology', None),
            'SerologyInvestigation': files.get('SerologyInvestigation', None),
            'OtherDocuments': files.get('OtherDocuments', None)})

        if serializer.is_valid():
            CaseNumber = serializer.validated_data.pop('CaseNumber')
            serializer.save(user=request.user)
            return Response({'status': 'success',
                            'message': 'data saved successfully'})

        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'}, status=400)


class ReportsPostForm(generics.GenericAPIView):
    serializer_class = ReportsSerializer
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        CaseNumber = request.data['CaseNumber'].replace('/', '')
        date = datetime.datetime.today().date()
        files = {}
        for field in ['OT_ProcedureSheets', 'AnaesthesiaNotes', 'OtherDocuments']:

            images = request.FILES.getlist(field)
            if images:
                for image in images:
                    if image.size > 400 * 1024:
                        return Response({'status': 'error', 'message': 'Each Image size should be less than 499 KB'}, status=400)
                pdf = PDFGenerator(images, CaseNumber, field)
                files[field] = SimpleUploadedFile(
                    f"{CaseNumber}_{date}_{field}.zip", pdf, content_type='application/zip')

        radiology_files = request.FILES.getlist('Radiology')
        if radiology_files:
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
                for file in radiology_files:
                    # Save the original file to disk temporarily
                    file_path = os.path.join(
                        os.path.dirname(file.name), file.name)
                    with open(file_path, 'wb+') as destination:
                        for chunk in file.chunks():
                            destination.write(chunk)
                    # Add the file to the ZIP archive
                    zip_file.write(file_path, os.path.basename(file_path))
                    # Remove the temporary file from disk
                    os.remove(file_path)
            # Create a SimpleUploadedFile object from the ZIP buffer
            zip_buffer.seek(0)
            files['Radiology'] = SimpleUploadedFile(f"{CaseNumber}_{date}_Radiology.zip",
                                                    zip_buffer.read(), content_type='application/zip')
        serializer = self.get_serializer(data={
            'CaseNumberID': request.data['CaseNumberID'],
            'CaseNumber': request.data['CaseNumber'],
            'Radiology': files.get('Radiology', None),
            'OT_ProcedureSheets': files.get('OT_ProcedureSheets', None),
            'AnaesthesiaNotes': files.get('AnaesthesiaNotes', None),
            'OtherDocuments': files.get('OtherDocuments', None),
        })
        if serializer.is_valid():
            CaseNumber = serializer.validated_data.pop('CaseNumber')
            serializer.save(user=request.user)
            return Response({'status': 'success',
                            'message': 'data saved successfully'})

        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'}, status=400)


class DischargesummuryPostAPI(generics.GenericAPIView):
    serializer_class = DischargeSummarySerialzer
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        CaseNumber = request.data['CaseNumber'].replace('/', '')
        date = datetime.datetime.today().date()
        files = {}
        for field in ['DischargeSummaryDocument', 'OtherDocuments']:

            images = request.FILES.getlist(field)
            if images:
                for image in images:
                    if image.size > 400 * 1024:
                        return Response({'status': 'error', 'message': 'Each Image size should be less than 499 KB'}, status=400)
                pdf = PDFGenerator(images, CaseNumber, field)
                files[field] = SimpleUploadedFile(
                    f"{CaseNumber}_{date}_{field}.zip", pdf, content_type='application/zip')

        serializer = self.get_serializer(data={
            'CaseNumberID': request.data['CaseNumberID'],
            'CaseNumber': request.data['CaseNumber'],
            'dischargeDate': request.data['dischargeDate'],
            'DischargeType': request.data['DischargeType'],
            'DischargeSummaryDocument': files.get('DischargeSummaryDocument', None),
            'OtherDocuments': files.get('OtherDocuments', None)
        })

        if serializer.is_valid():
            CaseNumber = serializer.validated_data.pop('CaseNumber')
            serializer.save(user=request.user)
            return Response({'status': 'success',
                            'message': 'data saved successfully'})

        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'}, status=400)


class DeathSummuryPostAPI(generics.GenericAPIView):
    serializer_class = DeathSummarySerialzer
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        CaseNumber = request.data['CaseNumber'].replace('/', '')
        date = datetime.datetime.today().date()
        files = {}
        for field in ['MortalityAudit', 'DeathCertificate', 'OtherDocuments']:

            images = request.FILES.getlist(field)
            if images:
                for image in images:
                    if image.size > 400 * 1024:
                        return Response({'status': 'error', 'message': 'Each Image size should be less than 499 KB'}, status=400)
                pdf = PDFGenerator(images, CaseNumber, field)
                files[field] = SimpleUploadedFile(
                    f"{CaseNumber}_{date}_{field}.zip", pdf, content_type='application/zip')

        serializer = self.get_serializer(data={
            'CaseNumberID': request.data['CaseNumberID'],
            'CaseNumber': request.data['CaseNumber'],
            'MortalityAudit': files.get('MortalityAudit', None),
            'DeathCertificate': files.get('DeathCertificate', None),
            'OtherDocuments': files.get('OtherDocuments', None)
        })

        if serializer.is_valid():
            CaseNumber = serializer.validated_data.pop('CaseNumber')
            serializer.save(user=request.user)
            return Response({'status': 'success',
                            'message': 'data saved successfully'})

        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'}, status=400)


class BloodDocumentsPostAPi(generics.GenericAPIView):
    serializer_class = BloodDocumentsSerializer
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        CaseNumber = request.data['CaseNumber'].replace('/', '')
        date = datetime.datetime.today().date()
        files = {}
        for field in ['BloodTransfusion', 'BTSticker', 'CrossMatchReport', 'OtherDocuments']:

            images = request.FILES.getlist(field)
            if images:
                for image in images:
                    if image.size > 400 * 1024:
                        return Response({'status': 'error', 'message': 'Each Image size should be less than 499 KB'}, status=400)
                pdf = PDFGenerator(images, CaseNumber, field)
                files[field] = SimpleUploadedFile(
                    f"{CaseNumber}_{date}_{field}.zip", pdf, content_type='application/zip')

        serializer = self.get_serializer(data={
            'CaseNumberID': request.data['CaseNumberID'],
            'CaseNumber': request.data['CaseNumber'],
            'BloodTransfusion': files.get('BloodTransfusion', None),
            'BTSticker': files.get('BTSticker', None),
            'CrossMatchReport': files.get('CrossMatchReport', None),
            'OtherDocuments': files.get('OtherDocuments', None),
        })

        if serializer.is_valid():
            CaseNumber = serializer.validated_data.pop('CaseNumber')
            serializer.save(user=request.user)
            return Response({'status': 'success',
                            'message': 'data saved successfully'})

        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'}, status=400)


# inserting the data in to database for .xls file
class DumpExcelInsertxls(generics.GenericAPIView):
    parser_classes = [MultiPartParser]
    serializer_class = DumpExcelSerializer

    def post(self, request, format=None):
        if 'excel_file' not in request.FILES:
            return Response({"status": "error", "message": "No file uploaded."}, status=400)

        excel_file = request.FILES["excel_file"]
        if excel_file.size == 0 or excel_file.name.endswith(".xls") != True:
            return Response({"status": "error", "message": "Only .xls file is supported."}, status=400)

        # excel_file = request.FILES["excel_file"]
        # if excel_file:
        wb = xlrd.open_workbook(file_contents=excel_file.read())
        sheet_name = wb.sheet_names()[0]
        worksheet = wb.sheet_by_name(sheet_name)
        data_list = []
        for row_idx in range(3, worksheet.nrows):
         
            row = worksheet.row_values(row_idx)
            data = DumpExcel(
                CaseNo=row[0], ClaimNo=row[1], RegistrationNumber=row[2], NHPMId=row[3],
                FamilyId=row[4], PatientDistrict=row[5], Gender=row[6], Age=row[7],
                CategoryDetails=row[8], ProcedureDetails=row[9], CaseType=row[10], CaseStatus=row[11], HospitalName=row[12],
                HospitalCode=row[13], HospitalDistrict=row[14], IPRegistrationDate=row[15], AdmissionDate=row[16], PreauthDate=row[17],
                PreauthAmount=row[18], PreauthApproveDate=row[19], PreauthApprovedAmount=row[20], PreauthRejectedDate=row[21],
                SurgeryDate=row[22], DeathDate=row[23], DischargeDate=row[24], ClaimSubmittedDate=row[25], ActualClaimSubmittedDate=row[26],
                ClaimInitaiatedAmount=row[27], CPDApprovedAmount=row[
                    28], ClaimApprovedAmount=row[29], AssignedFlag=row[30], AssignedUser=row[31],
                AssignedGroup=row[32], IPNumber=row[33], ActualRegistrationDate=row[34]
            )

            data_list.append(data)

        DumpExcel.objects.bulk_create(data_list)
        return Response({"status": "Success", "message": "Successfully Uploaded."})
        # else:
        #     return  Response({"status": "error", "Message": "excel_file can not be blank"} , status = 400)


# for inserting the data into database .xlsx file format


class DumpExcelInsertxlsx(generics.GenericAPIView):
    parser_classes = [MultiPartParser]
    serializer_class = DumpExcelSerializer

    def post(self, request, format=None):
        try:
            if 'excel_file' not in request.FILES:
                return Response({"status": "error", "message": "No file uploaded."}, status=400)

            excel_file = request.FILES["excel_file"]
            print(excel_file.name)
            if excel_file.size == 0 or excel_file.name.endswith(".xlsx") != True:
                return Response({"status": "error", 
                                 "message": "only .xlsx file is supported."},
                                   status=400)

            workbook = load_workbook(filename=excel_file)
            sheet_name = workbook.sheetnames[0]
            worksheet = workbook[sheet_name]
            data_list = [DumpExcel(*row) for row in worksheet.iter_rows(min_row=2, values_only=True)]

            DumpExcel.objects.bulk_create(data_list)
            return Response({"status": "Success",
                            "message": "Successfully Uploaded."})
        except:
            return Response({"status": "error",
                            "message": "Something Went Wrong Please Check your Excel File and upload Again."},
                            status=400)



# class MPClaimPiadPostApi(generics.GenericAPIView):
#     parser_classes = [MultiPartParser]
#     serializer_class = DumpExcelSerializer

#     def post(self, request, format=None):
#         if 'excel_file' not in request.FILES:
#             return Response({"status": "error", "message": "No file uploaded."}, status=400)

#         excel_file = request.FILES["excel_file"]
#         print(excel_file.name)
#         if excel_file.size == 0 or excel_file.name.endswith(".xlsx") != True:
#             return Response({"status": "error", "message": "only .xlsx file is supported."}, status=400)

#         workbook = load_workbook(filename=excel_file)
#         sheet_name = workbook.sheetnames[0]
#         worksheet = workbook[sheet_name]
#         data_list = []
#         for row in worksheet.iter_rows(min_row=2, values_only=True):
#             data = MPClaimPaidExcel(
#                 CaseNumber=row[0],WorkflowStatus=row[1],StatusUpdateDate=row[2],NHPMID=row[3],FamilyID=row[4],
#                 PatientName=row[5],Age=row[6],Gender=row[7],Patients_Home_District=row[8],
#                 Patients_Home_State=row[9],Patient_Phone_No=row[10],Address=row[11],Communication_Address=row[12],
#                 Communication_Contact_No=row[13],Communication_Village=row[14],Communication_Mandal=row[15],
#                 Communication_District=row[16],Communication_State=row[17],Speciality=row[18],Patient_IP_OP=row[19],Procedure_Code=row[20],
#                 Procedure=row[21],Procedure_Auto_Approve=row[22],Medical_or_Surgery=row[23],
#                 IP_Registration_Date=row[24],Hospital_Name=row[25],Admission_Date=row[26],PreAuth_Initiation_Date=row[27],PreAuth_Initiation_Amount=row[28],
#                 PreAuth_Cancel_Date=row[29],PreAuth_Approval_Date=row[30],PreAuth_Approval_Amount=row[31],PreAuth_Rejection_Date=row[32],
#                 Enhancement_Flag=row[33],Enhancement_Approved_Amount=row[34],Surgery_Date=row[35],
#                 Discharge_Date=row[36],Death_Date=row[37],Claim_Raised_Date=row[38],Claim_Paid_Amount=row[39],CPD_Approved_Date=row[40],CPD_rejected_Date=row[41],
#                 SHA_Approved_Date=row[42],Claim_Paid_Date=row[43],Claim_UTR_Number=row[44],Hospital_District=row[45],Hospital_State=row[46],
#                 PreAuth_approval_remarks=row[47],PreAuth_rejection_remarks=row[48],Claim_approval_remarks=row[49],Claim_rejection_remarks=row[50],
#                 Claim_Initiated_Amount=row[51],RF_Amount=row[52],TDS_Amount=row[53],Amount_paid_to_Hospital=row[54],Claim_Approved_Amount=row[55],
#                 Claim_Updated_Date=row[56],Preauth_Pending_Remarks=row[57],Preauth_Pending_Date=row[58],Preauth_Pending_Updated_Remarks=row[59],Preauth_Pending_Updated_Date=row[60],
#                 Claim_Pending_Remarks=row[61],Claim_Pending_Date=row[62],Claim_Pending_Updated_Remarks=row[63],Claim_Pending_Updated_Date=row[64],Last_Updated_User=row[65],
#                 Is_Aadhar_Benificaiary=row[66],BioAuth_at_Registration=row[67],BioAuth_at_Discharge=row[68],Erroneous_Initiated_Amount=row[69],
#                 Erroneous_Initiated_Date=row[70],Erroneous_Approved_Amount=row[71],Erroneous_Approved_Date=row[72],Erroneous_Paid_Date=row[73],
#                 Erroneous_UTR_Number=row[74],IP_Number=row[75],CPD_Pending_Count=row[76],CPD_Processing_Time=row[77],Revoked_Case=row[78],
#                 Revoked_Date=row[79],Revoked_Remarks=row[80],Insurance_Liable_Amount=row[81],Trust_Liable_Amount=row[82],
#                 Patient_Liable_Amount=row[83],Actual_Registration_Date=row[84],)

#             data_list.append(data)

#         MPClaimPaidExcel.objects.bulk_create(data_list)
#         return Response({"status": "Success", "message": "Successfully Uploaded."})


class MPClaimPiadPostApi(generics.GenericAPIView):
    parser_classes = [MultiPartParser]
    serializer_class = DumpExcelSerializer

    def post(self, request, format=None):
        excel_file = request.FILES.get("excel_file")
        if not excel_file:
            return Response(
                {"status": "error", "message": "No file uploaded."},
                status=400,
            )

        if not excel_file.name.endswith(".xlsx"):
            return Response(
                {"status": "error", 
                 "message": "Only .xlsx files are supported."},
                status=400,
            )

        workbook = load_workbook(filename=excel_file)
        worksheet = workbook[workbook.sheetnames[0]]
        data_list = [
            MPClaimPaidExcel(*row) for row in worksheet.iter_rows(min_row=2, values_only=True)
        ]
        print(data_list)
        MPClaimPaidExcel.objects.bulk_create(data_list)

        return Response(
            {"status": "Success", "message": "Successfully uploaded."},
            status=200,
        )


import csv
from io import StringIO


class DumpCSVInsert(generics.GenericAPIView):
    parser_classes = [MultiPartParser]
    serializer_class = DumpCSVSerializer

    def post(self, request, format=None):
        if 'csv_file' not in request.FILES:
            return Response({"status": "error", "message": "No file uploaded."}, status=400)

        csv_file = request.FILES["csv_file"]
        if csv_file.size == 0 or not csv_file.name.endswith(".csv"):
            return Response({"status": "error", "message": "Only .csv files are supported."}, status=400)

        # Convert the CSV file to a string and create a StringIO object
        csv_string = csv_file.read().decode('utf-8')
        csv_io = StringIO(csv_string)

        data_list = []
        csv_reader = csv.reader(csv_io)
        print(csv_reader)
        next(csv_reader) # Skip the header row

        # Loop through the rows and create DumpExcel objects
        for row in csv_reader:
            data = MPClaimPaidExcel(
                CaseNumber=row[0],WorkflowStatus=row[1],StatusUpdateDate=row[2],NHPMID=row[3],FamilyID=row[4],
                PatientName=row[5],Age=row[6],Gender=row[7],Patients_Home_District=row[8],
                Patients_Home_State=row[9],Patient_Phone_No=row[10],Address=row[11],Communication_Address=row[12],
                Communication_Contact_No=row[13],Communication_Village=row[14],Communication_Mandal=row[15],
                Communication_District=row[16],Communication_State=row[17],Speciality=row[18],Patient_IP_OP=row[19],Procedure_Code=row[20],
                Procedure=row[21],Procedure_Auto_Approve=row[22],Medical_or_Surgery=row[23],
                IP_Registration_Date=row[24],Hospital_Name=row[25],Admission_Date=row[26],PreAuth_Initiation_Date=row[27],PreAuth_Initiation_Amount=row[28],
                PreAuth_Cancel_Date=row[29],PreAuth_Approval_Date=row[30],PreAuth_Approval_Amount=row[31],PreAuth_Rejection_Date=row[32],
                Enhancement_Flag=row[33],Enhancement_Approved_Amount=row[34],Surgery_Date=row[35],
                Discharge_Date=row[36],Death_Date=row[37],Claim_Raised_Date=row[38],Claim_Paid_Amount=row[39],CPD_Approved_Date=row[40],CPD_rejected_Date=row[41],
                SHA_Approved_Date=row[42],Claim_Paid_Date=row[43],Claim_UTR_Number=row[44],Hospital_District=row[45],Hospital_State=row[46],
                PreAuth_approval_remarks=row[47],PreAuth_rejection_remarks=row[48],Claim_approval_remarks=row[49],Claim_rejection_remarks=row[50],
                Claim_Initiated_Amount=row[51],RF_Amount=row[52],TDS_Amount=row[53],Amount_paid_to_Hospital=row[54],Claim_Approved_Amount=row[55],
                Claim_Updated_Date=row[56],Preauth_Pending_Remarks=row[57],Preauth_Pending_Date=row[58],Preauth_Pending_Updated_Remarks=row[59],Preauth_Pending_Updated_Date=row[60],
                Claim_Pending_Remarks=row[61],Claim_Pending_Date=row[62],Claim_Pending_Updated_Remarks=row[63],Claim_Pending_Updated_Date=row[64],Last_Updated_User=row[65],
                Is_Aadhar_Benificaiary=row[66],BioAuth_at_Registration=row[67],BioAuth_at_Discharge=row[68],Erroneous_Initiated_Amount=row[69],
                Erroneous_Initiated_Date=row[70],Erroneous_Approved_Amount=row[71],Erroneous_Approved_Date=row[72],Erroneous_Paid_Date=row[73],
                Erroneous_UTR_Number=row[74],IP_Number=row[75],CPD_Pending_Count=row[76],CPD_Processing_Time=row[77],Revoked_Case=row[78],
                Revoked_Date=row[79],Revoked_Remarks=row[80],Insurance_Liable_Amount=row[81],Trust_Liable_Amount=row[82],
                Patient_Liable_Amount=row[83],Actual_Registration_Date=row[84],)
            

            data_list.append(data)

        # Use bulk_create to insert all the data into the database
        MPClaimPaidExcel.objects.bulk_create(data_list)

        return Response({"status": "Success", "message": "Successfully Uploaded."})
