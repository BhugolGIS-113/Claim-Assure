from openpyxl.styles import Font
import xlrd
from openpyxl.utils import get_column_letter
from itertools import islice
import openpyxl
from rest_framework import generics
from .serializers import *
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from io import BytesIO
from reportlab.pdfgen import canvas
from zipfile import ZipFile
from django.http import HttpResponse
from reportlab.lib.utils import ImageReader
from .models import *
from django.core.files.uploadedfile import SimpleUploadedFile
from PIL import Image
import os
from ClaimAssurance import settings
import datetime
import zipfile
from rest_framework.views import APIView
from django.shortcuts import get_object_or_404
import pandas as pd
import io
from rest_framework import status


def PDFGenerator(images, PreAuthID, name):
    buffer_list = []
    max_pdf_size = 400000  # 450 KB
    current_pdf_size = 0
    current_pdf_images = 0
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer)
    file_counter = 1

    for image in images:
        img = Image.open(image)
        img_file = BytesIO()
        if img.format == "JPEG":
            img.save(img_file, 'JPEG', optimize=True)
        elif img.format == "PNG":
            img.save(img_file, 'PNG', optimize=True)
        elif img.format == "JPG":
            img.save(img_file, 'JPG', optimize=True)
        img_file_size = img_file.tell()

        if current_pdf_size + img_file_size > max_pdf_size:
            pdf.save()
            buffer_list.append(buffer)
            buffer = BytesIO()
            pdf = canvas.Canvas(buffer)
            current_pdf_size = 0
            current_pdf_images = 0

        img_file.seek(0)
        image_reader = ImageReader(img_file)
        pdf.setPageSize((img.width, img.height))
        pdf.drawImage(image_reader, 0, 0)
        pdf.showPage()
        current_pdf_size += img_file_size
        current_pdf_images += 1

    pdf.save()
    buffer_list.append(buffer)
    zip_buffer = BytesIO()
    date = datetime.datetime.today().date()
    with ZipFile(zip_buffer, 'w') as zip_file:
        for buffer in buffer_list:
            buffer.seek(0)
            zip_file.writestr('{}_{}_{}_{}.pdf'.format(
                PreAuthID, date, name,  file_counter), buffer.read())
            file_counter += 1

    zip_buffer.seek(0)
    return zip_buffer.read()


class PreAuthFormView(generics.GenericAPIView):
    serializer_class = CombinedSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request):
        serializer = CombinedSerializer(data=request.data)
        if serializer.is_valid():
            validated_data = serializer.validated_data

            personal_info_serializer = PersonalInfoSerializer(data={
                'user': request.user.id,
                'NHPMID': validated_data.get('NHPMID'),
                'nameOfPatient': validated_data.get('nameOfPatient'),
                'adharNumber': validated_data.get('adharNumber'),
                'adharPhotograph': validated_data.get('adharPhotograph'),
                'DOB': validated_data.get('DOB'),
                'gender': validated_data.get('gender'),
                'mobileNumber': validated_data.get('mobileNumber'),
                'alternativeNumber': validated_data.get('alternativeNumber'),
                'addressLine1': validated_data.get('addressLine1'),
                'addressLine2': validated_data.get('addressLine2'),
                'district': validated_data.get('district'),
                'pincode': validated_data.get('pincode'),
                'RationCardNumber': validated_data.get('RationCardNumber'),
                'RationCardPhotograph': validated_data.get('RationCardPhotograph')})

            if personal_info_serializer.is_valid():
                personal_info = personal_info_serializer.save()
            
            else:
                key , value = list(personal_info_serializer.errors.items())[0]
                error_message = key+" ,"+ value[0]
                return Response({'status': 'error', 'message': error_message})

            # Genrating Sequential Preauth ID
            Last_ID = PreAuthDocument.objects.order_by('PreAuthID').last()
            if Last_ID:
                PreAuthID = 'PRE{:05}'.format(
                    int(Last_ID.PreAuthID[3:]) + 1)
            else:
                PreAuthID = 'PRE00001'

            date = datetime.datetime.today().date()
            files = {}
            for field in ['justification', 'on_BedPhotograph', 'admitCaseSheet', 'labReport']:
                images = request.FILES.getlist(field)

                if images:
                    for image in images:
                        name = image.name
                        if image.size > 400 * 1024:
                            return Response({'status': 'error', 'message': 'Each Image size should be less than 499 KB - {name}'.format(name=name)}, status=400)
                    pdf = PDFGenerator(images, PreAuthID, field)
                    files[field] = SimpleUploadedFile(
                        f"{PreAuthID}_{date}_{field}.zip", pdf, content_type='application/zip'
                    )
            radiology_files = request.FILES.getlist('radiologyReport')
            if radiology_files:
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
                    for file in radiology_files:
                        # Save the original file to disk temporarily
                        file_path = os.path.join(
                            os.path.dirname(file.name), file.name)
                        with open(file_path, 'wb+') as destination:
                            for chunk in file.chunks():
                                destination.write(chunk)
                        # Add the file to the ZIP archive
                        zip_file.write(file_path, os.path.basename(file_path))
                        # Remove the temporary file from disk
                        os.remove(file_path)
                # Create a SimpleUploadedFile object from the ZIP buffer
                zip_buffer.seek(0)
                files['radiologyReport'] = SimpleUploadedFile(f"{PreAuthID}_{date}_Radiology.zip",
                                                              zip_buffer.read() , content_type='application/zip')

            preauth_document_serializer = PreAuthDocumentSerializer(data={
                'user': request.user.id,
                'PreAuthID': PreAuthID,
                'PersonalInfoID': personal_info.id,
                'dateOfAdmission': validated_data.get('dateOfAdmission'),
                'dateOfPreAuth': validated_data.get('dateOfPreAuth'),
                'hospitalName': validated_data.get('hospitalName'),
                'hospitalCode': validated_data.get('hospitalCode'),
                'justification': files.get('justification', None),
                'on_BedPhotograph': files.get('on_BedPhotograph', None),
                'admitCaseSheet': files.get('admitCaseSheet', None),
                'labReport': files.get('labReport', None),
                'radiologyReport': files.get('radiologyReport', None), })

            if preauth_document_serializer.is_valid():
                preauth_document_serializer.save()
                return Response({'status': 'success', 'message': 'Data Saved Successfully'})
            else:

                key, value =list(preauth_document_serializer.errors.items())[0]
                error_message = key+" ,"+value[0]
                return Response({'status': 'error', 'message': error_message} , status=400)
        else:
            key, value =list(serializer.errors.items())[0]
            error_message = key+" ,"+value[0]
            return Response({'status': 'error',
                             'message': error_message} , status=400)


class FilterbyNHPMID(generics.GenericAPIView):
    serializer_class = PersonalInfoSerializer

    def get(self, request, NHPMID):
        PersonalInformation = PersonalInfo.objects.filter(
            NHPMID=NHPMID).first()

        if not PersonalInformation:
            return Response({'status': 'error',
                             'message': f'{NHPMID} ID is not found'}, status=400)
        PreAuthData = PreAuthDocument.objects.filter(
            PersonalInfoID=PersonalInformation.id)

        PreAuthDataserializer = PreAuthSearchDocumentSerializer(
            PreAuthData, many=True).data
        serializer = PersonalInfoSerializer(PersonalInformation).data
        return Response({'status': 'success',
                        'message': 'Data fetched successfully',
                         'personalInfo': serializer,
                         'PreAuthData': PreAuthDataserializer})


class SearchFilterbyPreAuthID(generics.ListAPIView):
    serializer_class = PreAuthSearchDocumentSerializer

    def get(self, request, PreAuthID):
        PreAuthData = PreAuthDocument.objects.filter(
            PreAuthID=PreAuthID)

        if not PreAuthData:
            return Response({'status': 'error',
                             'message': f'{PreAuthID} ID is not found'}, status=400)

        PersonalInformation = PersonalInfo.objects.filter(
            id=PreAuthData[0].PersonalInfoID.id)
        serializer = PersonalInfoSerializer(PersonalInformation[0]).data
        PreAuthDataserializer = PreAuthDocumentSerializer(
            PreAuthData, many=True).data

        return Response({'status': 'success',
                        'message': 'Data fetched successfully',
                         'personalInfo': serializer,
                         'PreAuthData': PreAuthDataserializer}, status=200)


class SearchFilterbyCaseNumber(generics.GenericAPIView):
    serializer_class = PreAuthSearchDocumentSerializer

    def get(self, request, CaseNumber, **kwargs):
        preauth_documents = PreAuthDocument.objects.filter(
            PreAuth_PreAuthID__CaseNumber=CaseNumber)

        if not preauth_documents:
            return Response({'status': 'error',
                             'message': f'There is no preauth ID available for - {CaseNumber}'},
                            status=400)
        CaseNo = PreAuthLinkCaseNumber.objects.get(CaseNumber=CaseNumber)
        CaseNo_Serializer = PreAuthLinkCaseNumberSerialzier(CaseNo).data
        PersonalInfo = preauth_documents[0].PersonalInfoID
        persnolInfo_serializer = PersonalInfoSerializer(PersonalInfo).data
        serilaizer = PreAuthSearchDocumentSerializer(preauth_documents[0]).data

        return Response({
            'status': 'success',
            'message': 'data fetched successfully',
            'casenumber_detail': CaseNo_Serializer,
            'persnolInfo': persnolInfo_serializer,
            'preAuth': serilaizer})


class ExistingPreAuthFormView(generics.GenericAPIView):
    serializer_class = ExistingPreAuthDocumentSerializer
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = ExistingPreAuthDocumentSerializer(data=request.data)
        if serializer.is_valid(raise_exception=True):
            data = serializer.validated_data
            Last_ID = PreAuthDocument.objects.order_by('PreAuthID').last()
            if Last_ID:
                PreAuthID = 'PRE{:05}'.format(
                    int(Last_ID.PreAuthID[3:]) + 1)
            else:
                PreAuthID = 'PRE00001'

            # Extract files from request and generate PDFs
            files = {}
            for field in ['justification', 'on_BedPhotograph', 'admitCaseSheet', 'labReport']:
                images = request.FILES.getlist(field)
                if images:
                    for image in images:
                        name = image.name
                        if image.size > 400 * 1024:
                            return Response({'status': 'error', 'message': 'Each Image size should be less than 499 KB - {name}'.format(name=name)}, status=400)
                    pdf = PDFGenerator(images, PreAuthID, field)
                    files[field] = SimpleUploadedFile(
                        f"{PreAuthID}_{datetime.datetime.today().date()}_{field}.zip",
                        pdf,
                        content_type='application/zip'
                    )

            radiology_files = request.FILES.getlist('Radiology')
            if radiology_files:
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
                    for file in radiology_files:
                        # Save the original file to disk temporarily
                        file_path = os.path.join(
                            os.path.dirname(file.name), file.name)
                        with open(file_path, 'wb+') as destination:
                            for chunk in file.chunks():
                                destination.write(chunk)
                        # Add the file to the ZIP archive
                        zip_file.write(file_path, os.path.basename(file_path))
                        # Remove the temporary file from disk
                        os.remove(file_path)
                # Create a SimpleUploadedFile object from the ZIP buffer
                zip_buffer.seek(0)
                files['Radiology'] = SimpleUploadedFile(f"{PreAuthID}_{datetime.datetime.today().date()}_Radiology.zip",
                                                        zip_buffer.read(),
                                                        content_type='application/zip')

            # Save pre-authorization document
            preauth_document_serializer = ExistingPreAuthserializer(data={
                'PreAuthID': PreAuthID,
                'PersonalInfoID': data.get('PersonalInfoID'),
                'dateOfAdmission': data.get('dateOfAdmission'),
                'dateOfPreAuth': data.get('dateOfPreAuth'),
                'hospitalName': data.get('hospitalName'),
                'hospitalCode': data.get('hospitalCode'),
                'justification': files.get('justification', None),
                'on_BedPhotograph': files.get('on_BedPhotograph', None),
                'admitCaseSheet': files.get('admitCaseSheet', None),
                'labReport': files.get('labReport', None),
                'radiologyReport': files.get('radiologyReport', None),

            })
            if not preauth_document_serializer.is_valid(raise_exception=True):
                key, value = list(
                    preauth_document_serializer.errors.items())[0]
                error_message = key+" , "+value[0]
                return Response({'status': 'error',
                                 'message': error_message})

            preauth_document_serializer.save(user=request.user)
            return Response({'status': 'success',
                            'message': 'Data Saved Successfully'})
        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({
                'status': 'error',
                'message': error_message})


class UpdatePreAuthDocumentView(generics.UpdateAPIView):
    serializer_class = PreAuthSearchDocumentUpdateSerializer
    parser_classes = [MultiPartParser]
    permission_classes = [IsAuthenticated]

    def update(self, request, PreAuthID, **kwargs):
        try:
            instance = PreAuthDocument.objects.get(PreAuthID = PreAuthID)
        except PreAuthDocument.DoesNotExist:
            return Response({'status': 'error',
                             'message' : 'PreAuth ID is not found'}, status=400)

        serializer = PreAuthSearchDocumentUpdateSerializer(
            instance, data=request.data, partial=True)
        if serializer.is_valid():
            validated_data = serializer.validated_data

            files = {}
            date = datetime.datetime.today().date()
            for key in ['justification', 'on_BedPhotograph', 'admitCaseSheet', 'labReport']:
                images = request.FILES.getlist(key)
                if images:
                    for image in images:
                        name = image.name
                        if image.size > 400 * 1024:
                            return Response({'status': 'error',
                                              'message': 'Each Image size should be less than 499 KB - {name}'.format(name=name)}, status=400)

                    zip_file = SimpleUploadedFile(f'{PreAuthID}_{date}_{key}.zip', PDFGenerator(
                        images, PreAuthID, key), content_type='application/zip')
                    files[key] = zip_file
                else:
                    files[key] = None

            radiology_files = request.FILES.getlist('radiologyReport')
            if radiology_files:
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
                    for file in radiology_files:
                        # Save the original file to disk temporarily
                        file_path = os.path.join(
                            os.path.dirname(file.name), file.name)
                        with open(file_path, 'wb+') as destination:
                            for chunk in file.chunks():
                                destination.write(chunk)
                        # Add the file to the ZIP archive
                        zip_file.write(file_path, os.path.basename(file_path))
                        # Remove the temporary file from disk
                        os.remove(file_path)
                # Create a SimpleUploadedFile object from the ZIP buffer
                zip_buffer.seek(0)
                files['radiologyReport'] = SimpleUploadedFile(f"{PreAuthID}_{date}_radiologyReport.zip",
                                                              zip_buffer.read(),
                                                              content_type='application/zip')

            preauth_document_serializer = preauth_document_serializer_update_serializer(instance, data={
                'dateOfAdmission': validated_data.get('dateOfAdmission'),
                'dateOfPreAuth': validated_data.get('dateOfPreAuth'),
                'hospitalName': validated_data.get('hospitalName'),
                'hospitalCode': validated_data.get('hospitalCode'),
                **files}, partial=True)

            if preauth_document_serializer.is_valid():
                preauth_document_serializer.save(user=request.user)
                return Response({'status': 'success', 'message': 'Data Updated Successfully'}, status=201)
            else:
                key, value = list(
                    preauth_document_serializer.errors.items())[0]
                error_message = key + ", " + value[0]
                return Response({'status': 'error', 'message': error_message}, status=400)
        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key + ", " + value[0]
            return Response({'status': 'error', 'message': error_message}, status=400)


class DeletePreAUth(APIView):
	permission_classes=[IsAuthenticated]
	def delete(self,request,PreAuthID,*args,**kwargs):
		try:
			instance=PreAuthDocument.objects.get(PreAuthID=PreAuthID)
		except PreAuthDocument.DoesNotExist as e:
			return Response({
				"message":"PreAuth ID Is not Found",
				"status":"error" 
				}, status=400)
		instance.delete()

		return Response({
			"message":"PreAuth data deleted successfully",
			"status":"success"
			},status=status.HTTP_200_OK)




class DownloadPreAuthZipFile(APIView):
    def get(self, request, PreAuthID):
        try:
            pre_auth_document = PreAuthDocument.objects.filter(
                PreAuthID=PreAuthID).latest('date_modified')
           
        except PreAuthDocument.DoesNotExist:
            return Response({'status': 'error', 'message': 'PreAuth ID not Found'}, status=400)

        date = datetime.datetime.today().date()
        response = HttpResponse(content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="{name}_{date}_PreAuth_Documents.zip"'.format(
            name=PreAuthID, date=date)
        zip_file = zipfile.ZipFile(response, 'w', zipfile.ZIP_DEFLATED)

        for field_name in ['justification', 'on_BedPhotograph', 'admitCaseSheet', 'labReport', 'radiologyReport']:
            field = getattr(pre_auth_document, field_name)
            if field and field.file:
                with zipfile.ZipFile(field.file, 'r') as zip_ref:
                    for zip_info in zip_ref.infolist():
                        zip_info.filename = os.path.basename(zip_info.filename)
                        zip_file.writestr(zip_info, zip_ref.read(zip_info.filename))

        zip_file.close()
        return response



class LinkingCaseNumberView(generics.GenericAPIView):
    serializer_class = CaseNumberLinkingSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request):
        data = request.data
        serializer = self.get_serializer(data=data)
        if serializer.is_valid():
            serializer.save(user=request.user)
            return Response({'message': 'Data Saved Successfully',
                            'status': 'success'})
        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'})

# This is for .CSV files 

# class ExcelMergeAPIView(generics.GenericAPIView):
#     parser_classes = [MultiPartParser]
#     serializer_class = ExcelFileSerializer

#     def post(self, request):
#         serializer = ExcelFileSerializer(data=request.data)
#         if serializer.is_valid():
#             # Load the CSV files
#             csv_files = request.FILES.getlist('file')
#             if len(csv_files) < 2:
#                 return Response({'message': 'At least two CSV files are required for Merging.',
#                                  'status': 'error'},
#                                 status=status.HTTP_400_BAD_REQUEST)

#             # Merge the CSV files
#             dfs = [pd.read_csv(f, header=2) for f in csv_files]
#             merged_df = pd.concat(dfs, ignore_index=True)

#             # Write the merged data to an XLSX file
#             output_file = io.BytesIO()
#             writer = pd.ExcelWriter(output_file, engine='openpyxl')
#             merged_df.to_excel(writer, index=False, sheet_name='Sheet1')

#             # Add the header only once
#             worksheet = writer.sheets['Sheet1']
#             for i, col in enumerate(merged_df.columns):
#                 worksheet.cell(row=1, column=i+1, value=col)

#             # Save the XLSX file
#             writer.save()
#             output_file.seek(0)
#             date = datetime.datetime.today().date()
#             # Return the merged file as a response
#             response = HttpResponse(output_file.read(),
#                                     content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#             response['Content-Disposition'] = 'attachment; filename= MERGE_DUMP_{date}.xlsx'.format(
#                 date=date)
#             return response


class ClaimFormView(generics.GenericAPIView):
    parser_classes = [MultiPartParser, FormParser]
    serializer_class = ClaimFormSerializer
    permission_classes = [IsAuthenticated]

    def post(self, request):
        CaseNumber = request.data['CaseNumber'].replace('/', '')
        date = datetime.datetime.today().date()
        files = {}
        for field in ['AdmitCase_ClinicalSheet', 'ICP', 'MedicationChart', 'InitialSheet', 'TPRChart',
                      'VitalChart', 'LOS_StayChart', 'CaseSheets_OtheDocuments', 'Microbiology',
                      'Biochemistry', 'Pathology', 'SerologyInvestigation', 'OT_ProcedureSheets',
                      'AnaesthesiaNotes', 'DischargeSummaryDocument', 'MortalityAudit', 'DeathCertificate',
                      'BloodTransfusion', 'BTSticker', 'CrossMatchReport', 'OtherDocuments']:

            images = request.FILES.getlist(field)
            if images:
                for image in images:
                    if image.size > 400 * 1024:
                        return Response({'status': 'error', 'message': 'Each Image size should be less than 499 KB'}, status=400)
                pdf = PDFGenerator(images, CaseNumber, field)
                files[field] = SimpleUploadedFile(
                    f"{CaseNumber}_{date}_{field}.zip", pdf, content_type='application/zip')

        # Zip the Radiology files
        radiology_files = request.FILES.getlist('Radiology')
        if radiology_files:
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
                for file in radiology_files:
                    # Save the original file to disk temporarily
                    file_path = os.path.join(
                        os.path.dirname(file.name), file.name)
                    with open(file_path, 'wb+') as destination:
                        for chunk in file.chunks():
                            destination.write(chunk)
                    # Add the file to the ZIP archive
                    zip_file.write(file_path, os.path.basename(file_path))
                    # Remove the temporary file from disk
                    os.remove(file_path)
            # Create a SimpleUploadedFile object from the ZIP buffer
            zip_buffer.seek(0)
            files['Radiology'] = SimpleUploadedFile(f"{CaseNumber}_{date}_Radiology.zip",
                                                    zip_buffer.read(), content_type='application/zip')

        serializer = ClaimFormSerializer(data={
            'CaseNumberID': request.data['CaseNumberID'],
            'CaseNumber': request.data['CaseNumber'],
            'dischargeDate': request.data['dischargeDate'],
            'DischargeType': request.data['DischargeType'],
            'Status': request.data['Status'],

            'AdmitCase_ClinicalSheet': files.get('AdmitCase_ClinicalSheet', None),
            'ICP': files.get('ICP', None),
            'MedicationChart': files.get('MedicationChart', None),
            'InitialSheet': files.get('InitialSheet', None),
            'TPRChart': files.get('TPRChart', None),
            'VitalChart': files.get('VitalChart', None),
            'LOS_StayChart': files.get('LOS_StayChart', None),
            'CaseSheets_OtheDocuments': files.get('CaseSheets_OtheDocuments', None),

            'Microbiology': files.get('Microbiology', None),
            'Biochemistry': files.get('Biochemistry', None),
            'Pathology': files.get('Pathology', None),
            'SerologyInvestigation': files.get('SerologyInvestigation', None),

            'Radiology': files.get('Radiology', None),
            'OT_ProcedureSheets': files.get('OT_ProcedureSheets', None),
            'AnaesthesiaNotes': files.get('AnaesthesiaNotes', None),
            'DischargeSummaryDocument': files.get('DischargeSummaryDocument', None),
            'MortalityAudit': files.get('MortalityAudit', None),
            'DeathCertificate': files.get('DeathCertificate', None),

            'BloodTransfusion': files.get('BloodTransfusion', None),
            'BTSticker': files.get('BTSticker', None),
            'CrossMatchReport': files.get('CrossMatchReport', None),
            'OtherDocuments': files.get('OtherDocuments', None),
        })

        if serializer.is_valid():
            serializer.save(user=request.user)
            return Response({'status': 'success',
                            'message': 'data saved successfully'})

        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'}, status=400)
        


#this is for .xls excel files 

class ExcelMergeAPIView(generics.GenericAPIView):
    serializer_class = ExcelFileSerializer
    parser_classes = [MultiPartParser]

    def post(self, request, format=None):
        serializer = ExcelFileSerializer(data=request.data)
        if serializer.is_valid():
            # Load the Excel files
            excel_files = []
            for file in request.FILES.getlist('file'):
                excel_files.append(xlrd.open_workbook(
                    file_contents=file.read()))
            if len(excel_files) == 1:
                return Response({'status': 'error',
                                'message': "There is only one Excel file , To Merge the Excel's at least two or more files will be needed."})
            # Merge the Excel files

            merged_workbook = openpyxl.Workbook()
            merged_sheet = merged_workbook.active
            merged_sheet.title = 'Merged Sheet'
            header_added = False  # To keep track of whether the header has been added
            row_index_merged = 1  # To keep track of the row index in the merged sheet
            for excel_file in excel_files:
                for sheet in excel_file.sheets():
                    for row_index in range(sheet.nrows):
                        # If the sheet is the first sheet, and header has not been added yet
                        if not header_added and row_index == 2:
                            # Add the headers to the merged sheet
                            for col_index in range(sheet.ncols):
                                column_letter = get_column_letter(col_index+1)
                                merged_sheet[column_letter+str(row_index_merged)] = sheet.cell_value(
                                    row_index, col_index)
                                merged_sheet[column_letter +
                                             str(row_index_merged)].font = Font(bold=True)
                            row_index_merged += 1
                            header_added = True  # Set header_added to True so that it is not added again
                        elif row_index > 2:  # If the row is after the header row
                            for col_index in range(sheet.ncols):
                                column_letter = get_column_letter(col_index+1)
                                merged_sheet[column_letter+str(row_index_merged)] = sheet.cell_value(
                                    row_index, col_index)
                            row_index_merged += 1

            # Write the merged Excel file to a BytesIO object
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=MERGE_DUMP.xlsx'
            merged_workbook.save(response)

            return response
        else:
            key, value = list(serializer.errors.items())[0]
            error_message = key+" , "+value[0]
            return Response({'message': error_message,
                            'status': 'error'})

import xlrd

class DumpExcelInsert(generics.GenericAPIView):
    parser_classes = [MultiPartParser]
    serializer_class =DumpExcelSerializer


    def post(self, request, format=None):
        if 'excel_file' not in request.FILES:
            return Response({"status": "error", "message": "No file uploaded."} , status=400)

        excel_file = request.FILES["excel_file"]
        if excel_file.size == 0:
            return Response({"status": "error", "message": "File is empty."} , status=400)

        # excel_file = request.FILES["excel_file"]
        # if excel_file:
        wb = xlrd.open_workbook(file_contents=excel_file.read())
        sheet_name = wb.sheet_names()[0]
        worksheet = wb.sheet_by_name(sheet_name)
        data_list = []
        for row_idx in range(3, worksheet.nrows):
            row = worksheet.row_values(row_idx)
            data = DumpExcel(
                CaseNo=row[0], ClaimNo=row[1], RegistrationNumber=row[2], NHPMId=row[3], 
                FamilyId=row[4], PatientDistrict=row[5], Gender=row[6], Age=row[7], 
                CategoryDetails=row[8], ProcedureDetails=row[9], CaseType=row[10], CaseStatus=row[11], HospitalName=row[12], 
                HospitalCode=row[13], HospitalDistrict=row[14], IPRegistrationDate=row[15], AdmissionDate=row[16], PreauthDate=row[17], 
                PreauthAmount=row[18], PreauthApproveDate=row[19], PreauthApprovedAmount=row[20], PreauthRejectedDate=row[21], 
                SurgeryDate=row[22], DeathDate=row[23], DischargeDate=row[24], ClaimSubmittedDate=row[25], ActualClaimSubmittedDate=row[26], 
                ClaimInitaiatedAmount=row[27], CPDApprovedAmount=row[28], ClaimApprovedAmount=row[29], AssignedFlag=row[30], AssignedUser=row[31], 
                AssignedGroup=row[32], IPNumber=row[33], ActualRegistrationDate=row[34]
            )

            data_list.append(data)

        DumpExcel.objects.bulk_create(data_list)
        return Response({"status": "Success", "message": "Successfully Uploaded."})
        # else:
            # return  Response({"status": "error", "Message": "excel_file can not be blank"} , status = 400)
